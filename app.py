# app.py
# 服裝物料採購單自動化處理系統
# Step 1: 檔案上傳 + 合併儲存格(向下填充)+ 表格預覽
# Step 2: 雙層拆分(供應商 → 款號)+ 上傳記錄 sidebar + 清空按鈕
# Step 3: 真款號 vs 備註分類 + 智慧取價(優先序:成本价 > 大货价 > 单价 > 报价)
# Step 4A: 標準材料庫上傳 + 三階紅燈比對(🟢已驗證 / 🟡規格不同 / 🚨全新料)

import io
import re
import zipfile

import pandas as pd
import streamlit as st


# ── 頁面設定 ─────────────────────────────────────────────
st.set_page_config(
    page_title="物料採購單處理系統",
    page_icon="📋",
    layout="wide",
)


# ── Session state 初始化 ─────────────────────────────────
# upload_history:累積這次 session 所有上傳過的檔名(關閉瀏覽器才清)
# uploaded_blobs:對應的檔案二進位內容,讓 sidebar 可以提供下載
# uploader_key:清空按鈕按下時 +1,藉換 key 讓 file_uploader 重置
if "upload_history" not in st.session_state:
    st.session_state.upload_history = []
if "uploaded_blobs" not in st.session_state:
    st.session_state.uploaded_blobs = {}
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0
# Step 4A:標準庫資料(只存解析後的清單,不存原始 bytes)
if "std_library" not in st.session_state:
    st.session_state.std_library = None  # 解析後的 dict
if "std_library_filename" not in st.session_state:
    st.session_state.std_library_filename = None


# ── 左側 sidebar 渲染函式(在主流程結尾處呼叫,避免函式未定義錯誤)───
def render_sidebar():
    with st.sidebar:
        st.header("📜 本次上傳記錄")

        if st.session_state.upload_history:
            st.caption(f"共 {len(st.session_state.upload_history)} 筆(點擊下載原始檔)")
            for i, fname in enumerate(st.session_state.upload_history, 1):
                blob = st.session_state.uploaded_blobs.get(fname)
                if blob is not None:
                    st.download_button(
                        label=f"📥 {i}. {fname}",
                        data=blob,
                        file_name=fname,
                        mime="application/octet-stream",
                        key=f"dl_{i}_{fname}",
                        use_container_width=True,
                    )
                else:
                    st.markdown(f"`{i}.` {fname} _(內容已清)_")
        else:
            st.info("尚未上傳任何檔案")

        st.divider()

        # ── Step 4A:標準材料庫上傳區 ────────────────
        st.subheader("📚 標準材料庫")
        if st.session_state.std_library:
            st.success(f"✅ 已載入:{st.session_state.std_library_filename}")
            n_names = len(st.session_state.std_library["name_set"])
            st.caption(f"共 {n_names} 筆標準品名")
            if st.button("🗑️ 移除標準庫", use_container_width=True):
                st.session_state.std_library = None
                st.session_state.std_library_filename = None
                st.rerun()
        else:
            std_file = st.file_uploader(
                "上傳標準材料庫(供比對用)",
                type=["xlsx"],
                accept_multiple_files=False,
                help="不上傳也能用,但無法做紅燈警告",
                key="std_lib_uploader",
            )
            if std_file is not None:
                try:
                    with st.spinner("解析標準庫中(12MB 約需 30 秒)..."):
                        st.session_state.std_library = parse_standard_library(std_file.getvalue())
                        st.session_state.std_library_filename = std_file.name
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 無法解析:{e}")

        st.divider()

        # 「清空」只清當前 file_uploader,記錄與標準庫都保留
        if st.button("🔄 清空,重新開始", type="primary", use_container_width=True):
            st.session_state.uploader_key += 1
            st.rerun()


# 注意:render_sidebar() 故意延後到主流程「最尾端」呼叫,
# 因為 sidebar 內會用到 parse_standard_library() 等下方定義的函式。
# Python 從上往下讀,函式呼叫前必須先看過函式定義。


# ── 主畫面標題 ───────────────────────────────────────────
st.title("📋 服裝物料採購單自動化處理系統")
st.caption("Step 2:檔案上傳 → 合併儲存格 → 雙層拆分(供應商 → 款號)")


# ── 檔案上傳區 ────────────────────────────────────────────
uploaded_files = st.file_uploader(
    "請選擇要處理的檔案(可一次多選)",
    type=["xlsx", "xls", "csv"],
    accept_multiple_files=True,
    help="支援 .xlsx / .xls / .csv;Excel 多工作表會全部讀進來",
    key=f"file_uploader_{st.session_state.uploader_key}",
)

# 累積上傳記錄 + 暫存原始檔位元組(讓 sidebar 可下載)
# 若有新檔名加入,立刻 rerun 一次,讓 sidebar 看到最新記錄
if uploaded_files:
    new_names = []
    for f in uploaded_files:
        if f.name not in st.session_state.upload_history:
            new_names.append(f.name)
            # getvalue() 取出位元組,存進 session(只活在當前瀏覽器分頁)
            st.session_state.uploaded_blobs[f.name] = f.getvalue()
    if new_names:
        st.session_state.upload_history.extend(new_names)
        st.rerun()

if not uploaded_files:
    st.info("👆 請從上方上傳一個或多個檔案")
    render_sidebar()  # 主流程提早結束前先把 sidebar 畫出來
    st.stop()


# ── 讀檔函式(同 Step 1,Step 5 修正:備註欄不 ffill)──────
# 這些欄位**不應該** ffill — 它們的空白是「真實沒值」,不是合併儲存格留白
# 否則上一筆物料的備註會被錯誤地灌到下一筆物料上
_NO_FFILL_COL_KEYWORDS = ["备注", "備註", "報價備註", "报价备注", "remark", "Remark", "REMARK"]


def _ffill_with_exceptions(df: pd.DataFrame, header_row_idx: int = 10) -> pd.DataFrame:
    """
    對 df 做 ffill,但把指定關鍵字的欄位排除(不 ffill)。
    header_row_idx:用第幾行(0-indexed)當欄名來判斷該欄是不是要排除。
    """
    df_filled = df.ffill(axis=0)
    if df.shape[0] <= header_row_idx:
        return df_filled
    headers = df.iloc[header_row_idx]
    for col_idx, header in enumerate(headers):
        if isinstance(header, str) and any(kw in header for kw in _NO_FFILL_COL_KEYWORDS):
            # 該欄不 ffill,還原為原始 NaN 狀態
            df_filled.iloc[:, col_idx] = df.iloc[:, col_idx]
    return df_filled


def read_file(file) -> dict:
    """讀取上傳檔,回傳 {sheet_name: DataFrame}(已 ffill,但備註欄保留原始)。"""
    name = file.name.lower()
    sheets: dict[str, pd.DataFrame] = {}

    if name.endswith(".csv"):
        df = pd.read_csv(file, header=None, dtype=str)
        df = _ffill_with_exceptions(df)
        sheets["CSV"] = df
    else:
        xls = pd.ExcelFile(file, engine="openpyxl")
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=str)
            df = _ffill_with_exceptions(df)
            sheets[sheet_name] = df

    return sheets


# ── Step 2 核心邏輯:抽資料區 + 雙層拆分 ─────────────────

# 從樣本分析得知:真表頭固定在第 10 行(0-indexed)
# 第 0~9 行是公司抬頭區、供應商/聯絡人/客戶資訊
HEADER_ROW = 10

# 抬頭區的「供應商」備援值位置(raw_sample_1 沒有「供應商」欄時用)
SUPPLIER_LABEL_ROW = 5
SUPPLIER_VALUE_COL = 2


def parse_sheet(df: pd.DataFrame, sheet_name: str) -> tuple[pd.DataFrame, dict]:
    """
    把一張 ffilled 過的 sheet 拆成「資料區 + 抬頭資訊」。

    回傳:
      data_df: 資料區(以第 10 行為欄名,從第 11 行起)
      header_info: {default_supplier, sheet_name}
    """
    if df.shape[0] <= HEADER_ROW:
        return pd.DataFrame(), {"default_supplier": None, "sheet_name": sheet_name}

    # 取第 10 行當欄名,並處理「空白欄名」「重複欄名」問題
    # 否則 Streamlit 顯示表格會 ValueError(duplicate columns / NaN columns)
    raw_headers = df.iloc[HEADER_ROW].tolist()
    headers, seen = [], {}
    for i, h in enumerate(raw_headers):
        # NaN / 空白 → 用 _col_X 占位
        if h is None or pd.isna(h) or not str(h).strip() or str(h).lower() == "nan":
            name = f"_col_{i}"
        else:
            name = str(h).strip()
        # 重複名 → 加序號
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)

    data_df = df.iloc[HEADER_ROW + 1:].copy()
    data_df.columns = headers
    data_df = data_df.reset_index(drop=True)

    # 去掉完全空白的行(資料區尾巴常見)
    data_df = data_df.dropna(how="all").reset_index(drop=True)

    # 從抬頭區抓供應商備援值(raw_sample_1 用,因為資料區沒有供應商欄)
    default_supplier = None
    if df.shape[0] > SUPPLIER_LABEL_ROW and df.shape[1] > SUPPLIER_VALUE_COL:
        cell = df.iat[SUPPLIER_LABEL_ROW, SUPPLIER_VALUE_COL]
        if cell and str(cell).strip() and str(cell).lower() != "nan":
            default_supplier = str(cell).strip()

    return data_df, {"default_supplier": default_supplier, "sheet_name": sheet_name}


# ── Step 3 核心:真款號判別 + 備註分類 + 智慧取價 ─────────

# 黑名單(明顯不是款號的字串)
_FAKE_CODE_BLACKLIST = {"TOTAL", "注:", "注:", "備註:", "备注:", "—", "-"}


def is_real_style_code(value) -> bool:
    """
    判斷一個 cell 是否為「真款號」。
    真款號:短(< 20 字)、含數字、不在黑名單、不是「1./2./3.」這種編號條款。
    """
    if value is None or pd.isna(value):
        return False
    s = str(value).strip()
    if not s:
        return False
    if s.upper() in {x.upper() for x in _FAKE_CODE_BLACKLIST}:
        return False
    # 長度太長 → 多半是中文備註句
    if len(s) > 20:
        return False
    # 沒有任何數字 → 多半是標籤,不是編號
    if not any(c.isdigit() for c in s):
        return False
    # 開頭是「1.」「2、」這種編號條款
    if len(s) >= 2 and s[0].isdigit() and s[1] in ".．、 ":
        return False
    return True


def classify_data_rows(
    data_df: pd.DataFrame, style_col: str
) -> tuple[pd.DataFrame, list[str]]:
    """
    分離資料行為「真款號物料」與「備註行」。
    真款號 → real_df
    其他 → 取 style_col 的字串當作備註,集中為一個 list(去重保序)
    """
    if data_df.empty or style_col not in data_df.columns:
        return data_df.copy(), []

    real_mask = data_df[style_col].apply(is_real_style_code)
    real_df = data_df[real_mask].copy().reset_index(drop=True)
    notes_df = data_df[~real_mask]

    notes: list[str] = []
    seen: set[str] = set()
    for _, row in notes_df.iterrows():
        text = row[style_col]
        if text is None or pd.isna(text):
            continue
        text = str(text).strip()
        # 過濾無意義小標題、合計列
        if not text or text.upper() == "TOTAL" or text in {"注:", "注:"}:
            continue
        if text not in seen:
            seen.add(text)
            notes.append(text)

    return real_df, notes


# 智慧取價:依優先序找出「成本價」欄
# 順序:成本价 > 供应商成本价 > 大货价 > 单价 > 报价
_PRICE_PRIORITY = ["成本价", "供应商成本价", "大货价", "单价", "报价"]


def find_cost_column(columns) -> str | None:
    """從欄位名稱中找出最優先的價格欄;找不到回 None"""
    for keyword in _PRICE_PRIORITY:
        for col in columns:
            if isinstance(col, str) and keyword in col:
                return col
    return None


# ── Step 4A:標準材料庫讀取 + 三階紅燈比對 ─────────────────

# 標準庫每張 sheet 的設定:表頭行(0-indexed)、品名/料號欄、規格欄、跳過旗標
# 可比對的欄位 → key 值用標準化字串
_STD_SHEET_CONFIG = {
    "3F-Brax Zippers": {"header_row": 1, "name_keys": ["品名", "款号"], "size_keys": ["规格", "尺码"]},
    "Metal Trims":     {"header_row": 0, "name_keys": ["Article No.品名", "品名"], "size_keys": ["Size尺寸", "Size"]},
    "Rivet":           {"header_row": 1, "name_keys": ["Item", "FS article number"], "size_keys": []},
    "Non Metal Trims": {"header_row": 0, "name_keys": ["Article No.品名", "品名"], "size_keys": ["Size尺寸", "Size"]},
    "Label":           {"header_row": 0, "name_keys": ["Article No.", "品名"], "size_keys": ["Size", "尺寸"]},
    "Tape":            {"header_row": 0, "name_keys": ["Article No.品名", "品名"], "size_keys": ["Size尺寸", "Size"]},
    "Paper Tags":      {"skip": True},  # 設計圖,無結構化資料
}


def _strip_styles_xlsx(file_bytes: bytes) -> bytes:
    """
    繞過 openpyxl 3.1.5 對某些 .xlsx 的 styles.xml 解析 crash 問題。
    把 styles.xml 換成最小骨架後重新打包。資料完整保留,只丟掉樣式。
    """
    src = io.BytesIO(file_bytes)
    dst = io.BytesIO()
    minimal_styles = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        b'<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
        b'<borders count="1"><border/></borders>'
        b'<cellStyleXfs count="1"><xf/></cellStyleXfs>'
        b'<cellXfs count="1"><xf/></cellXfs>'
        b'</styleSheet>'
    )
    with zipfile.ZipFile(src, "r") as zin:
        with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if item == "xl/styles.xml":
                    data = minimal_styles
                zout.writestr(item, data)
    return dst.getvalue()


def normalize_for_compare(s) -> str:
    """字串正規化(供比對用):去全部空白、轉大寫、去除常見單位寫法差異。"""
    if s is None or pd.isna(s):
        return ""
    text = str(s).upper()
    # 去掉所有空白(含全形空白)
    text = re.sub(r"\s+", "", text)
    # 統一全形→半形數字/標點
    text = text.replace(",", ",").replace(":", ":")
    return text


def parse_standard_library(file_bytes: bytes) -> dict:
    """
    解析使用者上傳的標準庫 .xlsx,回傳:
    {
        'name_set': set[str],         # 所有「品名」正規化字串集合
        'name_size_set': set[tuple],  # (品名正規化, 規格正規化) 二元組集合
        'sheets_info': dict           # 每張 sheet 的統計
    }
    """
    safe_bytes = _strip_styles_xlsx(file_bytes)
    xls = pd.ExcelFile(io.BytesIO(safe_bytes), engine="openpyxl")

    name_set: set[str] = set()
    name_size_set: set[tuple] = set()
    sheets_info: dict = {}

    for sheet_name in xls.sheet_names:
        cfg = _STD_SHEET_CONFIG.get(sheet_name)
        if cfg is None:
            sheets_info[sheet_name] = {"status": "未知 sheet,已跳過", "count": 0}
            continue
        if cfg.get("skip"):
            sheets_info[sheet_name] = {"status": "跳過(無結構化資料)", "count": 0}
            continue

        df = pd.read_excel(xls, sheet_name=sheet_name, header=cfg["header_row"], dtype=str)

        # 找品名欄(用 name_keys 列表逐一嘗試,取第一個符合的)
        name_col = None
        for key in cfg["name_keys"]:
            for col in df.columns:
                if isinstance(col, str) and key in col:
                    name_col = col
                    break
            if name_col:
                break

        # 找規格欄
        size_col = None
        for key in cfg.get("size_keys", []):
            for col in df.columns:
                if isinstance(col, str) and key in col:
                    size_col = col
                    break
            if size_col:
                break

        if name_col is None:
            sheets_info[sheet_name] = {"status": "找不到品名欄,跳過", "count": 0}
            continue

        count = 0
        for _, row in df.iterrows():
            raw_name = row.get(name_col)
            n = normalize_for_compare(raw_name)
            if not n:
                continue
            name_set.add(n)
            if size_col:
                s = normalize_for_compare(row.get(size_col))
                if s:
                    name_size_set.add((n, s))
                else:
                    name_size_set.add((n, ""))
            else:
                name_size_set.add((n, ""))
            count += 1

        sheets_info[sheet_name] = {
            "status": f"✅ 已讀取(品名欄:{name_col} / 規格欄:{size_col or '無'})",
            "count": count,
        }

    return {
        "name_set": name_set,
        "name_size_set": name_size_set,
        "sheets_info": sheets_info,
    }


def _match_one(query: str, library: dict, size_query: str) -> str:
    """
    單一 key 的三階比對。回傳 🟢 / 🟡 / 🚨。
    寬鬆比對:正規化後雙向「包含」即視為命中。
    """
    if not query:
        return "🚨"
    matched_names = set()
    for n_lib in library["name_set"]:
        if not n_lib:
            continue
        if query in n_lib or n_lib in query:
            matched_names.add(n_lib)

    if not matched_names:
        return "🚨"

    for n_lib in matched_names:
        sizes = {s for (nm, s) in library["name_size_set"] if nm == n_lib}
        if "" in sizes:
            return "🟢"
        if not size_query:
            return "🟡"
        for s_lib in sizes:
            if size_query in s_lib or s_lib in size_query:
                return "🟢"
    return "🟡"


# 三階優先序:🟢 最好 → 🟡 → 🚨 最差
_RANK = {"🟢": 0, "🟡": 1, "🚨": 2, "⚪": 3}


# ── Step 5:模板填值匯出 ZIP ──────────────────────────────

# 公司抬頭(已固定為嘉善凯翔)— Step 5 採用此名稱
COMPANY_NAME = "嘉善凯翔服饰有限公司"

# target_format 模板路徑(放在 repo 內,部署時隨程式上雲)
TEMPLATE_PATH = "templates/target_format.xlsx"
TEMPLATE_SHEET = "辅料"  # 模板 sheet 名

# 模板抬頭區關鍵 cell:程式會把這幾格的值替換掉
TEMPLATE_HEADER_CELLS = {
    "supplier_cell": "C6",        # 供應商
    "date_cell": "K5",            # 訂料日期
    # 收貨人/電話/交貨期/交貨地點 raw 通常沒有,留空繼承
}

# 模板資料區設定
TEMPLATE_DATA_HEADER_ROW = 11   # 表頭第 11 行
TEMPLATE_DATA_START_ROW = 12    # 資料從第 12 行起
TEMPLATE_EXAMPLE_ROWS = 2       # 模板實際只有 row 12, 13 兩行範例物料(row 15 是合計列,不能動)

# raw 欄名 → target 欄名 的映射
# 兩邊欄名常見差異(全形空白、簡繁差異)做寬鬆比對
# 注意:raw 的「单价」欄常被填垃圾字串「单价」,刻意排除 → 改用 cost_col 強制映射
# 「金额」欄通常是 raw 公式算的結果,我們改用 target 的公式自動算 → 排除
_RAW_TO_TARGET_MAP = {
    "款号": "款号",
    "品名": "品名",
    "客户编号": "客户编号",
    "颜色": "颜色",
    "成品数": "成品数",
    "规格": "规格cm",
    "规格cm": "规格cm",
    "尺码": "规格cm",
    "尺寸": "规格cm",
    "单耗": "单耗",
    "单位": "单位",
    "总订料数": "总订料数",
    "订料数量": "总订料数",
    "产前样数量": "产前样数量",
    "备注": "备注",
}

# 排除清單:這些 raw 欄名直接跳過,不寫入 target
_RAW_SKIP_COLS = {"单价", "金额"}


def _normalize_col_name(s):
    """把欄名正規化:去空白、繁簡統一(粗略,只去常見全形空白)"""
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).strip()


def build_supplier_excel(
    supplier_name: str,
    styles_dict: dict,
    sheet_notes: list,
    cost_col: str | None,
    template_bytes: bytes,
    order_date: str = "",
) -> bytes:
    """
    為單一供應商產生一份 Excel 檔(每個款號一張 sheet)。

    參數:
      supplier_name : 供應商名稱(會填到 C6)
      styles_dict   : { 款號: 物料 DataFrame }
      sheet_notes   : 該 sheet 的備註(填到每個物料的「备注」欄末尾)
      cost_col      : raw 中的成本價欄名(會映射到 target 的「单价」)
      template_bytes: target_format.xlsx 的二進位
      order_date    : 訂料日期字串

    回傳:
      bytes:Excel 檔的二進位
    """
    import openpyxl
    from openpyxl.cell.cell import MergedCell

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    template_ws = wb[TEMPLATE_SHEET]

    # 對每個款號:複製模板 sheet 並重新命名
    first = True
    for style_code, style_df in styles_dict.items():
        if first:
            ws = template_ws
            ws.title = str(style_code)[:30]  # Excel sheet 名上限 31 字
            first = False
        else:
            ws = wb.copy_worksheet(template_ws)
            ws.title = str(style_code)[:30]

        # 1. 替換抬頭區
        ws[TEMPLATE_HEADER_CELLS["supplier_cell"]] = supplier_name
        if order_date:
            ws[TEMPLATE_HEADER_CELLS["date_cell"]] = order_date

        # 2. 清掉模板範例物料區的值(只清 row 12, 13 — row 15 黃色合計列保留)
        for r in range(TEMPLATE_DATA_START_ROW, TEMPLATE_DATA_START_ROW + TEMPLATE_EXAMPLE_ROWS):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                cell.value = None

        # 2.5 物料筆數超過模板範例(2 筆)時,要先 insert_rows 讓黃色合計列往下推
        n_materials = len(style_df)
        if n_materials > TEMPLATE_EXAMPLE_ROWS:
            from copy import copy
            from openpyxl.utils import get_column_letter

            extra = n_materials - TEMPLATE_EXAMPLE_ROWS
            insert_at = TEMPLATE_DATA_START_ROW + TEMPLATE_EXAMPLE_ROWS  # row 14

            # ── insert_rows 前先抓出所有「會被推下去」的合併儲存格,事後手動修正 ──
            # openpyxl insert_rows 不會自動更新合併範圍,必須自己重設
            merges_to_shift = []
            for merge in list(ws.merged_cells.ranges):
                if merge.min_row >= insert_at:
                    merges_to_shift.append((
                        merge.min_row, merge.max_row,
                        merge.min_col, merge.max_col,
                    ))
                    ws.unmerge_cells(str(merge))

            ws.insert_rows(idx=insert_at, amount=extra)

            # 重新建立合併範圍(每個範圍都往下推 extra 行)
            for min_r, max_r, min_c, max_c in merges_to_shift:
                new_range = (
                    f"{get_column_letter(min_c)}{min_r + extra}:"
                    f"{get_column_letter(max_c)}{max_r + extra}"
                )
                ws.merge_cells(new_range)

            # 手動從 row 13(最後一個範例)複製樣式 + 公式到新插入的行
            template_row = TEMPLATE_DATA_START_ROW + 1  # row 13
            for offset in range(extra):
                new_row = insert_at + offset
                for c in range(1, ws.max_column + 1):
                    src = ws.cell(row=template_row, column=c)
                    dst = ws.cell(row=new_row, column=c)
                    if isinstance(dst, MergedCell):
                        continue
                    if src.has_style:
                        dst.font = copy(src.font)
                        dst.fill = copy(src.fill)
                        dst.border = copy(src.border)
                        dst.alignment = copy(src.alignment)
                        dst.number_format = src.number_format
                    if isinstance(src.value, str) and src.value.startswith("="):
                        old_ref = str(template_row)
                        new_ref = str(new_row)
                        new_formula = re.sub(
                            rf"(?<![0-9]){old_ref}(?![0-9])",
                            new_ref,
                            src.value,
                        )
                        dst.value = new_formula

        # 3. 從 target 第 11 行讀出表頭欄名 → 對應 column index
        target_headers = {}  # {欄名: column_idx}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=TEMPLATE_DATA_HEADER_ROW, column=c).value
            if v:
                target_headers[_normalize_col_name(v)] = c

        # 4. 建立 raw 欄名 → target column_idx 的映射
        col_mapping = {}  # {raw_col_name: target_col_idx}
        for raw_col in style_df.columns:
            raw_norm = _normalize_col_name(raw_col)
            # 跳過排除清單(单价/金额 — 改用 cost_col 與公式)
            if raw_norm in _RAW_SKIP_COLS:
                continue
            # 直接同名匹配
            if raw_norm in target_headers:
                col_mapping[raw_col] = target_headers[raw_norm]
                continue
            # 走 mapping 表
            for raw_key, target_key in _RAW_TO_TARGET_MAP.items():
                if raw_key in raw_norm:
                    target_norm = _normalize_col_name(target_key)
                    if target_norm in target_headers:
                        col_mapping[raw_col] = target_headers[target_norm]
                        break
        # cost_col 強制映射到 target 的「单价」(放在最後,優先級最高)
        if cost_col and "单价" in target_headers:
            col_mapping[cost_col] = target_headers["单价"]

        # 5. 填入物料資料(從 row 12 起)
        # 「备注」欄只放 raw 該物料自己的備註(由 col_mapping 自動帶),不再塞 sheet 級備註
        # sheet 級備註(出口规定、1./2./3. 條款)保留在 target 模板的固定位置(row 16-19)
        for i, (_, row) in enumerate(style_df.iterrows()):
            target_row = TEMPLATE_DATA_START_ROW + i
            for raw_col, target_col_idx in col_mapping.items():
                val = row[raw_col]
                if pd.notna(val) and str(val).strip():
                    cell = ws.cell(row=target_row, column=target_col_idx)
                    if isinstance(cell, MergedCell):
                        continue
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue
                    cell.value = val

    # 存成 bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_zip_from_supplier_files(files: dict) -> bytes:
    """
    files: { filename: bytes_content }
    回傳:zip 檔的二進位
    """
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, content in files.items():
            zf.writestr(fname, content)
    return out.getvalue()


@st.cache_data
def load_template_bytes() -> bytes:
    """讀取打包在 repo 中的 target_format.xlsx 模板"""
    with open(TEMPLATE_PATH, "rb") as f:
        return f.read()


def compare_with_library(name, code, size, library: dict) -> str:
    """
    雙 key 比對:品名 OR 客戶編號 任一命中即算找到。
    取兩個 key 比對結果中「最好」的那一個。

    🟢 任一命中 + 規格相符
    🟡 任一命中但規格不同
    🚨 兩個都找不到
    ⚪ 沒上傳標準庫 / 兩個都是空字串
    """
    if not library:
        return "⚪"

    n_query = normalize_for_compare(name)
    c_query = normalize_for_compare(code)
    s_query = normalize_for_compare(size)

    if not n_query and not c_query:
        return "⚪"

    results = []
    if n_query:
        results.append(_match_one(n_query, library, s_query))
    if c_query:
        results.append(_match_one(c_query, library, s_query))

    # 取最好的結果(rank 最小)
    best = min(results, key=lambda r: _RANK.get(r, 9))
    return best


def split_by_supplier_and_style(data_df: pd.DataFrame, header_info: dict) -> dict:
    """
    雙層拆分:先按供應商,再按款號。
    回傳:{ supplier: { style: sub_df } }
    """
    if data_df.empty:
        return {}

    # 找「供应商」欄(優先用資料區的;raw_sample_2 有這欄)
    supplier_col = None
    for col in data_df.columns:
        if isinstance(col, str) and "供应商" in col:
            supplier_col = col
            break

    # 找「款号」欄
    style_col = None
    for col in data_df.columns:
        if isinstance(col, str) and "款号" in col:
            style_col = col
            break

    if style_col is None:
        return {}

    # 第一層:按供應商分組
    if supplier_col:
        # 資料區有「供应商」欄
        groups_by_supplier = {}
        for s, g in data_df.groupby(supplier_col, dropna=False):
            if pd.isna(s) or not str(s).strip():
                continue
            groups_by_supplier[str(s).strip()] = g.copy()
    else:
        # 資料區沒有,退而用抬頭區的供應商;再退用 sheet 名
        default = (
            header_info.get("default_supplier")
            or header_info.get("sheet_name")
            or "未知供應商"
        )
        groups_by_supplier = {default: data_df.copy()}

    # 第二層:每個供應商裡按款號分組
    result = {}
    for supplier, supplier_df in groups_by_supplier.items():
        styles = {}
        for style, style_df in supplier_df.groupby(style_col, dropna=False):
            if pd.isna(style) or not str(style).strip():
                continue
            styles[str(style).strip()] = style_df.copy()
        if styles:
            result[supplier] = styles

    return result


# ── Step 5:匯出資料收集器 ──────────────────────────────
# 結構:{供應商名: {"styles": {款號: df}, "notes": [...], "cost_col": str}}
# 同一供應商跨多份檔/多 sheet 的款號會合併
export_collector: dict = {}


# ── 逐檔解析並展示 ────────────────────────────────────────
for file in uploaded_files:
    st.divider()
    st.subheader(f"📄 {file.name}")

    try:
        sheets = read_file(file)
    except Exception as e:
        st.error(f"❌ 讀取失敗:{e}")
        continue

    # 總體統計
    total_rows = sum(len(df) for df in sheets.values())
    total_nan = sum(int(df.isna().sum().sum()) for df in sheets.values())
    st.caption(
        f"工作表數:{len(sheets)}  /  總行數:{total_rows}  /  "
        f"殘留 NaN 格:{total_nan}"
    )

    # 對每張 sheet 做「原始預覽 + 拆分結果」
    for sheet_name, df in sheets.items():
        with st.expander(
            f"📑 工作表:{sheet_name}  —  {df.shape[0]} 行 × {df.shape[1]} 欄",
            expanded=True,
        ):
            # ① 原始 ffilled 預覽(收合在 expander 內,預設展開)
            st.markdown("##### 1️⃣ 原始預覽(ffill 後)")
            st.dataframe(df, use_container_width=True, hide_index=True, height=250)

            # ② 雙層拆分結果(Step 3 強化:識別假款號為備註 + 智慧取價)
            st.markdown("##### 2️⃣ 雙層拆分(供應商 → 款號)")

            data_df, header_info = parse_sheet(df, sheet_name)

            if data_df.empty:
                st.warning(
                    f"⚠️ 此工作表資料區為空(可能總行數不足 {HEADER_ROW + 1} 行)"
                )
                continue

            # 找款號欄(在 split 之前先找出來,以便分類)
            style_col_for_class = next(
                (c for c in data_df.columns if isinstance(c, str) and "款号" in c),
                None,
            )
            if style_col_for_class is None:
                st.warning("⚠️ 找不到「款号」欄,無法處理")
                continue

            # ── Step 3:分離真款號物料 vs 備註 ──
            real_df, sheet_notes = classify_data_rows(data_df, style_col_for_class)

            # ── Step 3:智慧取價 ──
            cost_col = find_cost_column(real_df.columns)

            # 顯示 sheet 級的備註(出口規定、條款等)
            if sheet_notes:
                with st.container(border=True):
                    st.markdown(f"📌 **此 sheet 共 {len(sheet_notes)} 條備註**(會套用到此 sheet 所有款號)")
                    for note in sheet_notes:
                        st.markdown(f"- {note}")

            # 顯示成本價欄判斷結果
            if cost_col:
                st.success(f"💰 自動抓到成本價欄:**{cost_col}**(優先序:成本价 > 大货价 > 单价 > 报价)")
            else:
                st.info("💰 未找到成本價相關欄位")

            # 用真款號 df 做拆分
            split_result = split_by_supplier_and_style(real_df, header_info)

            if not split_result:
                st.warning("⚠️ 過濾掉假款號後,沒有任何真款號可拆分")
                continue

            n_suppliers = len(split_result)
            n_styles = sum(len(v) for v in split_result.values())
            st.caption(
                f"📊 拆出 **{n_suppliers}** 個供應商,共 **{n_styles}** 個真款號"
                f"(已過濾 {len(sheet_notes)} 條備註)"
            )

            # ── Step 5:把這份資料累積到 export_collector ──
            for supplier, styles in split_result.items():
                if supplier not in export_collector:
                    export_collector[supplier] = {
                        "styles": {},
                        "notes": [],
                        "cost_col": cost_col,
                    }
                # 同款號累積資料(若已存在則合併)
                for style_code, sdf in styles.items():
                    if style_code in export_collector[supplier]["styles"]:
                        export_collector[supplier]["styles"][style_code] = pd.concat(
                            [export_collector[supplier]["styles"][style_code], sdf],
                            ignore_index=True,
                        )
                    else:
                        export_collector[supplier]["styles"][style_code] = sdf.copy()
                # 累積備註
                for note in sheet_notes:
                    if note not in export_collector[supplier]["notes"]:
                        export_collector[supplier]["notes"].append(note)

            # 第一層 tabs:供應商
            supplier_tabs = st.tabs([f"🏭 {s}" for s in split_result.keys()])
            for stab, (supplier, styles) in zip(supplier_tabs, split_result.items()):
                with stab:
                    # 第二層:每個款號一個 expander
                    for style, sdf in styles.items():
                        # ── Step 4A:逐筆物料做紅燈比對 ──
                        # 找此筆款的「品名」「客戶編號」「規格」欄
                        name_col_for_compare = next(
                            (c for c in sdf.columns if isinstance(c, str) and "品名" in c),
                            None,
                        )
                        code_col_for_compare = next(
                            (c for c in sdf.columns
                             if isinstance(c, str) and ("客户编号" in c or "客戶編號" in c or "Article" in c)),
                            None,
                        )
                        size_col_for_compare = next(
                            (c for c in sdf.columns
                             if isinstance(c, str) and ("规格" in c or "尺码" in c or "尺寸" in c)),
                            None,
                        )
                        # 算出比對狀態 — 雙 key 比對:品名 OR 客戶編號任一命中即綠
                        statuses = []
                        if name_col_for_compare or code_col_for_compare:
                            for _, mat_row in sdf.iterrows():
                                name = mat_row.get(name_col_for_compare) if name_col_for_compare else None
                                code = mat_row.get(code_col_for_compare) if code_col_for_compare else None
                                size = mat_row.get(size_col_for_compare) if size_col_for_compare else None
                                statuses.append(
                                    compare_with_library(name, code, size, st.session_state.std_library)
                                )
                        else:
                            statuses = ["⚪"] * len(sdf)

                        # 統計
                        n_green = statuses.count("🟢")
                        n_yellow = statuses.count("🟡")
                        n_red = statuses.count("🚨")
                        n_white = statuses.count("⚪")
                        # 款號標題若有紅燈,加警告 emoji
                        title_warn = " 🚨" if n_red > 0 else (" 🟡" if n_yellow > 0 else "")

                        with st.expander(
                            f"🏷️ 款號:{style}  ({len(sdf)} 筆物料){title_warn}",
                            expanded=(n_red > 0),  # 有紅燈自動展開
                        ):
                            # 顯示這筆款號套用了哪些備註(目前是 sheet 級)
                            if sheet_notes:
                                st.caption(f"📌 套用備註({len(sheet_notes)} 條)")
                            # 顯示成本價欄(若有)
                            if cost_col and cost_col in sdf.columns:
                                cost_values = sdf[cost_col].dropna().unique().tolist()
                                if cost_values:
                                    st.caption(
                                        f"💰 此款成本價(欄「{cost_col}」):"
                                        f"{', '.join(map(str, cost_values[:5]))}"
                                    )
                            # 比對統計
                            if st.session_state.std_library:
                                st.caption(
                                    f"🚦 比對結果: 🟢 {n_green} 筆 / 🟡 {n_yellow} 筆 / "
                                    f"🚨 {n_red} 筆 / ⚪ {n_white} 筆未比對"
                                )
                            else:
                                st.caption("⚪ 尚未上傳標準庫,無法比對(👈 在左側 sidebar 上傳)")

                            # 加上「比對狀態」欄到表格(放第一欄)
                            sdf_display = sdf.copy()
                            sdf_display.insert(0, "🚦", statuses)

                            st.dataframe(
                                sdf_display,
                                use_container_width=True,
                                hide_index=True,
                                height=220,
                            )


# ── Step 5:匯出 ZIP 區塊 ──────────────────────────────
if export_collector:
    st.divider()
    st.header("📦 匯出整理後的採購單")

    n_suppliers_total = len(export_collector)
    n_styles_total = sum(len(v["styles"]) for v in export_collector.values())
    st.markdown(
        f"準備匯出 **{n_suppliers_total}** 個供應商檔,共 **{n_styles_total}** 個款號 sheet"
    )

    # 顯示即將匯出的清單
    with st.expander("👀 預覽即將匯出的檔案結構", expanded=False):
        for sup, info in export_collector.items():
            st.markdown(f"📄 **採購單_{sup}.xlsx**")
            for style_code in info["styles"].keys():
                n_mat = len(info["styles"][style_code])
                st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;📑 sheet:`{style_code}` ({n_mat} 筆物料)")

    # 紅燈確認鎖(Step 4B 的承諾)
    # 計算所有物料的紅燈總數
    n_red_total = 0
    for sup, info in export_collector.items():
        for style_code, sdf in info["styles"].items():
            name_c = next((c for c in sdf.columns if isinstance(c, str) and "品名" in c), None)
            code_c = next((c for c in sdf.columns if isinstance(c, str) and ("客户编号" in c or "Article" in c)), None)
            size_c = next((c for c in sdf.columns if isinstance(c, str) and ("规格" in c or "尺码" in c or "尺寸" in c)), None)
            for _, row in sdf.iterrows():
                nm = row.get(name_c) if name_c else None
                cd = row.get(code_c) if code_c else None
                sz = row.get(size_c) if size_c else None
                if compare_with_library(nm, cd, sz, st.session_state.std_library) == "🚨":
                    n_red_total += 1

    # 紅燈確認
    can_export = True
    if n_red_total > 0:
        st.warning(
            f"⚠️ 偵測到 **{n_red_total} 筆紅燈物料**(標準庫找不到)\n\n"
            "依公司規定,紅燈物料需要主管確認。請勾選下方確認框後才能匯出。"
        )
        can_export = st.checkbox(
            f"✅ 我已確認這 {n_red_total} 筆紅燈物料皆已通過主管/相關人員審核",
            key="red_light_confirmed",
        )
    elif st.session_state.std_library is None:
        st.info("💡 提示:左側 sidebar 上傳標準材料庫即可啟用紅燈警告。目前未上傳,直接匯出。")

    # 產生按鈕(產生完把 bytes 暫存到 session_state,讓多個下載按鈕共用)
    if st.button(
        "🚀 產生匯出檔案",
        type="primary",
        disabled=not can_export,
        use_container_width=True,
    ):
        try:
            with st.spinner("正在產生供應商 Excel 檔案..."):
                template_bytes = load_template_bytes()
                supplier_files = {}
                for sup_name, info in export_collector.items():
                    excel_bytes = build_supplier_excel(
                        supplier_name=sup_name,
                        styles_dict=info["styles"],
                        sheet_notes=info["notes"],
                        cost_col=info["cost_col"],
                        template_bytes=template_bytes,
                    )
                    safe_name = re.sub(r'[\\/:*?"<>|]', "_", sup_name)
                    supplier_files[f"採購單_{safe_name}.xlsx"] = excel_bytes
                zip_bytes = build_zip_from_supplier_files(supplier_files)
            # 暫存到 session_state(讓重整不會消失)
            st.session_state["export_files"] = supplier_files
            st.session_state["export_zip"] = zip_bytes
            st.success(f"✅ 已產生 {len(supplier_files)} 份檔案,請選擇下載方式")
        except Exception as e:
            st.error(f"❌ 產生失敗:{e}")
            import traceback
            st.code(traceback.format_exc())

    # 顯示下載按鈕(產生過後一直留著)
    if st.session_state.get("export_files"):
        files = st.session_state["export_files"]
        zip_bytes = st.session_state["export_zip"]

        st.markdown("---")
        st.markdown("#### 📥 下載選項")

        # 整包 ZIP
        st.download_button(
            label=f"📦 下載全部({len(files)} 個檔)— ZIP 打包",
            data=zip_bytes,
            file_name="採購單匯出.zip",
            mime="application/zip",
            use_container_width=True,
            type="primary",
        )

        # 個別下載
        st.markdown("**或單獨下載某一個供應商檔:**")
        # 排成兩欄,讓畫面比較不那麼擠
        cols = st.columns(2)
        for i, (fname, fbytes) in enumerate(files.items()):
            with cols[i % 2]:
                st.download_button(
                    label=f"📄 {fname}",
                    data=fbytes,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"single_dl_{fname}",
                    use_container_width=True,
                )


# ── 主流程結束,最後渲染 sidebar(此時所有函式都已定義完成)──────
render_sidebar()
